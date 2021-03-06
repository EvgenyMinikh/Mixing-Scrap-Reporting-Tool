from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.uic import *
from configparser import ConfigParser
from json import load
from os import path, sep
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPDF
from openpyxl import load_workbook
from re import findall, sub
from subprocess import run
from datetime import datetime

this_script_dir = path.dirname(path.realpath(__file__))

NUMBER_OF_COPIES_DEFAULT_VALUE = 1
PDF_BROWSER_PATH = this_script_dir + '\\SumatraPDF.exe'
MAIN_UI_WINDOW_PATH = this_script_dir + '\\main_window.ui'
CONFIG_FILENAME = this_script_dir + '\\config.cfg'
SHIFT_SUPERVISORS_FILE_PATH = this_script_dir + '\\shift_supervisors.csv'
OPERATORS_FILE_PATH = this_script_dir + '\\operators.csv'
INCONSISTENCY_REASONS_FILE_PATH = this_script_dir + '\\list_items.json'
PDF_TIMESTAMP_FORMAT = '\\image_%d%m%y_%H%M%S.pdf'

config = ConfigParser()
config.read_file(open(CONFIG_FILENAME))
BUFFER_WORKBOOK_PATH = config.get('Paths', 'BUFFER_WORKBOOK_PATH')
IS_WB_HAS_MACROS = True if BUFFER_WORKBOOK_PATH.split('.')[-1].lower() == "xlsm" else False
FIRST_LINE_NUMBER = int(config.get('List Settings', 'FIRST_LINE_NUMBER'))
LAST_LINE_NUMBER = int(config.get('List Settings', 'LAST_LINE_NUMBER'))
SHIFTS = config.get('List Settings', 'SHIFTS')
DATE_FORMAT = config.get('Common Config', 'DATE_FORMAT')
SHEET_NAME = config.get('Paths', 'SHEET_NAME')
LABEL_OUT_PATH_PDF = config.get('Paths', 'LABEL_OUT_PATH_PDF')
LABEL_TEMPLATE_PATH = config.get('Paths', 'LABEL_TEMPLATE_PATH')
LABEL_OUT_PATH_SVG = config.get('Paths', 'LABEL_OUT_PATH_SVG')


def get_mixing_line_numbers_list(first, last):
    result = [str(i) for i in range(first, last + 1)]
    result.insert(0, "")
    return result


def get_shift_letters(st):
    result = [s for s in st]
    result.insert(0, "")
    return result


def get_user_list_from_file(file_path):
    result = []
    with open(file_path, mode='r', encoding='UTF8') as f:
        for line in f:
            result.append(line.strip())

    result.insert(0, "")
    return sorted(result)


def read_json_from_file(path):
    with open(path) as json_file:
        json_data = load(json_file)

    return json_data


def create_new_SVG_file_with_data(template_path, output_path, data_values):
    column_names_limits = set_column_width_limit(get_excel_cells_order())

    with open(template_path, mode='r', encoding='UTF8') as f:
        content = f.read()
        for key, value in data_values.items():
            value = str(value)

            if column_names_limits[key] != -1:
                list_values = findall(r'.{1,%d}' % column_names_limits[key], value)

                for i in range(len(list_values)):
                    value = list_values[i]
                    content = content.replace('{%s%d}' % (key, i), value)
            else:
                content = content.replace('{%s}' % key, value)

    with open(output_path, mode='w', encoding='UTF8') as f:

        for st in column_names_limits:
            content = sub(r'\{%s\d+\}' % st, '', content)  # Clean up SVG template from redundant fillers

        f.write(content)


def get_inconsistency_types_list(json_data):
    value_list = sorted(json_data.keys())
    return value_list


def data_checker(data_values):
    result = []
    for key, value in data_values.items():
        if value == '':
            result.append('Поле "{}" должно быть заполнено'.format(key))

    if len(result) > 0:
        return '\n'.join(result)

    return ''


def get_excel_cells_order():
    column_names = dict()
    column_names["Дата производства"] = 'A'
    column_names["Линия"] = 'B'
    column_names["Наименование продукции"] = 'C'
    column_names["Смена производства"] = 'D'
    column_names["Смена списания"] = 'E'
    column_names["Оператор 1"] = 'F'
    column_names["Оператор 2"] = 'G'
    column_names["Тип несоответствия"] = 'H'
    column_names["Причина несоответствия"] = 'I'
    column_names["Комментарии"] = 'J'
    column_names["Обнаружил"] = 'K'
    column_names["Бланк оформил"] = 'L'
    return column_names


def set_column_width_limit(column_names):
    column_names_limits = dict()
    for key in column_names.keys():
        column_names_limits[key] = -1

    column_names_limits["Дата списания"] = -1
    column_names_limits["Комментарии"] = 45
    column_names_limits["Наименование продукции"] = 35

    return column_names_limits


def check_workbook_ready_to_write(wb_path):
    file_name = wb_path.split(sep)[-1]
    full_wb_temp_path = wb_path.replace(file_name, '~$' + file_name)
    is_file_exists = path.isfile(full_wb_temp_path)

    return is_file_exists


def write_data_into_workbook(wb_path, data_to_write, aux_sheet):

    def write_data_into_ws(wb, data_to_write, sheet_name):
        column_names = get_excel_cells_order()
        ws = wb[sheet_name]
        line_number_to_write = len(ws['A']) + 1

        for key in data_to_write.keys():
            if key in column_names.keys():
                cell_address = column_names[key] + str(line_number_to_write)
                ws[cell_address] = data_to_write[key]

    wb = load_workbook(filename=wb_path, keep_vba=IS_WB_HAS_MACROS)
    write_data_into_ws(wb, data_to_write, SHEET_NAME)

    if aux_sheet not in wb.sheetnames:
        wb.create_sheet(aux_sheet)

    write_data_into_ws(wb, data_to_write, aux_sheet)
    wb.save(wb_path)


def get_pdf_name_to_save():
    return LABEL_OUT_PATH_PDF + datetime.now().strftime(PDF_TIMESTAMP_FORMAT)


class MainWindow(QMainWindow):
    def __init__(self, *args, **kwargs):
        super(MainWindow, self).__init__(*args, **kwargs)
        loadUi(MAIN_UI_WINDOW_PATH, self)
        self.set_current_date()

        # fill comboboxes with data
        self.comboBox_line_number.addItems(get_mixing_line_numbers_list(FIRST_LINE_NUMBER, LAST_LINE_NUMBER))
        self.comboBox_shift_number.addItems(get_shift_letters(SHIFTS))
        self.comboBox_writeoff_shift_number.addItems(get_shift_letters(SHIFTS))
        self.comboBox_blank_author.addItems(get_user_list_from_file(SHIFT_SUPERVISORS_FILE_PATH))
        self.comboBox_operator1.addItems(get_user_list_from_file(OPERATORS_FILE_PATH))
        self.comboBox_explorer.addItems(get_user_list_from_file(OPERATORS_FILE_PATH))

        self.operator2_list = get_user_list_from_file(OPERATORS_FILE_PATH)
        self.operator2_list[0] = 'Нет'
        self.comboBox_operator2.addItems(self.operator2_list)

        json_data = read_json_from_file(INCONSISTENCY_REASONS_FILE_PATH)
        self.comboBox_inconsistency_type.addItems(get_inconsistency_types_list(json_data))
        self.comboBox_inconsistency_type.activated.connect(lambda: self.get_inconsistency_reason_list(json_data))

        # Assign actions for buttons
        self.pushButton_ClearFields.clicked.connect(self.clean_all_forms)
        self.pushButton_Print.clicked.connect(self.print_label)
        self.pushButton_Saveto_PDF.clicked.connect(self.save_label)

        self.setFixedSize(self.size())
        self.show()

    # fill comboboxes with current date
    def set_current_date(self):
        current_date = QDate.currentDate()
        self.dateEdit_producing_date.setDate(current_date)
        self.dateEdit_writeoff_date.setDate(current_date)

    def clean_all_forms(self):
        self.set_current_date()
        self.comboBox_line_number.setCurrentIndex(0)
        self.comboBox_shift_number.setCurrentIndex(0)
        self.comboBox_writeoff_shift_number.setCurrentIndex(0)
        self.comboBox_blank_author.setCurrentIndex(0)
        self.comboBox_operator1.setCurrentIndex(0)
        self.comboBox_operator2.setCurrentIndex(0)
        self.comboBox_explorer.setCurrentIndex(0)
        self.spinBox_number_of_copies.setValue(NUMBER_OF_COPIES_DEFAULT_VALUE)
        self.comboBox_inconsistency_type.setCurrentIndex(0)
        self.comboBox_inconsistency_reason.setCurrentIndex(0)
        self.lineEdit_product_name.clear()
        self.lineEdit_comments.clear()

    def get_all_values_from_forms(self):
        data_values = dict()
        data_values["Дата производства"] = (self.dateEdit_producing_date.date()).toPyDate().strftime(DATE_FORMAT)
        data_values["Линия"] = self.comboBox_line_number.currentText()
        data_values["Наименование продукции"] = self.lineEdit_product_name.text()
        data_values["Смена производства"] = self.comboBox_shift_number.currentText()
        data_values["Смена списания"] = self.comboBox_writeoff_shift_number.currentText()
        data_values["Оператор 1"] = self.comboBox_operator1.currentText()
        data_values["Оператор 2"] = self.comboBox_operator2.currentText()
        data_values["Тип несоответствия"] = self.comboBox_inconsistency_type.currentText()
        data_values["Причина несоответствия"] = self.comboBox_inconsistency_reason.currentText()
        data_values["Комментарии"] = self.lineEdit_comments.text()
        data_values["Обнаружил"] = self.comboBox_explorer.currentText()
        data_values["Бланк оформил"] = self.comboBox_blank_author.currentText()
        data_values["Дата списания"] = (self.dateEdit_writeoff_date.date()).toPyDate().strftime(DATE_FORMAT)
        return data_values

    def file_is_opened_dialog_window(self):
        msg = "Не могу записать данные в файл. Файл {}\nкем-то открыт в Excel.\n\nДля продолжения работы сначала закройте файл".format(BUFFER_WORKBOOK_PATH)
        self.plainTextEdit_StatusField.clear()
        self.plainTextEdit_StatusField.insertPlainText(msg)
        QMessageBox.warning(self, "Can not save data", msg, buttons=QMessageBox.Ok)

    def print_label(self):
        if check_workbook_ready_to_write(BUFFER_WORKBOOK_PATH):
            self.file_is_opened_dialog_window()
            return

        data_values = self.get_all_values_from_forms()
        data_check_message = data_checker(data_values)
        self.plainTextEdit_StatusField.insertPlainText(data_check_message)

        if data_check_message == '':
            write_data_into_workbook(BUFFER_WORKBOOK_PATH, data_values, data_values["Тип несоответствия"])
            create_new_SVG_file_with_data(LABEL_TEMPLATE_PATH, LABEL_OUT_PATH_SVG, data_values)
            self.plainTextEdit_StatusField.clear()
            self.clean_all_forms()

            pdf_filename_to_save = get_pdf_name_to_save()

            self.plainTextEdit_StatusField.insertPlainText('Ярлык сохранен в {} и отправлен на печать'.format(pdf_filename_to_save))
            drawing = svg2rlg(LABEL_OUT_PATH_SVG)
            renderPDF.drawToFile(drawing, pdf_filename_to_save)
            run([PDF_BROWSER_PATH, '-print-dialog', '-exit-when-done', pdf_filename_to_save])

    def save_label(self):
        if check_workbook_ready_to_write(BUFFER_WORKBOOK_PATH):
            self.file_is_opened_dialog_window()
            return

        data_values = self.get_all_values_from_forms()
        data_check_message = data_checker(data_values)
        self.plainTextEdit_StatusField.insertPlainText(data_check_message)

        if data_check_message == '':
            write_data_into_workbook(BUFFER_WORKBOOK_PATH, data_values, data_values["Тип несоответствия"])
            create_new_SVG_file_with_data(LABEL_TEMPLATE_PATH, LABEL_OUT_PATH_SVG, data_values)
            self.plainTextEdit_StatusField.clear()
            self.clean_all_forms()

            pdf_filename_to_save = get_pdf_name_to_save()

            self.plainTextEdit_StatusField.insertPlainText('Ярлык сохранен в {}'.format(pdf_filename_to_save))
            drawing = svg2rlg(LABEL_OUT_PATH_SVG)
            renderPDF.drawToFile(drawing, pdf_filename_to_save)

    def get_inconsistency_reason_list(self, json_data):
        value_list = json_data[self.comboBox_inconsistency_type.currentText()]
        self.comboBox_inconsistency_reason.clear()
        self.comboBox_inconsistency_reason.addItems(value_list)

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Quit', "Закрыть окно?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()


if __name__ == '__main__':
    app = QApplication([])
    main_window = MainWindow()
    app.exec_()